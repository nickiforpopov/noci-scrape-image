import logging
import requests, json,re,os,datetime
import imghdr
from bs4 import BeautifulSoup

from openpyxl import load_workbook

from LR import LR
#from updates import update_label
import urllib.parse
from PIL import Image
from PIL import JpegImagePlugin 
JpegImagePlugin._getmp = lambda: None   

def decode_url(url):
    pattern = r'\\u[0-9a-fA-F]{4}'
    matches = re.findall(pattern, url)
    for match in matches:
        decoded = bytes(match.encode('ascii')).decode('unicode_escape')
        url = url.replace(match, r'' + decoded)
    decoded_url = urllib.parse.unquote(url)
    return decoded_url


###ADD IF ELSE IF WORKS TRUE ELSE FALSE
def getsku(filepath,skuRow):    

    #anchor = search.anchor

    filepath = filepath.replace('"','')
    print(filepath)

    global searchIds
    searchIds = []
    workbook = load_workbook(filepath)
    print(filepath)
    global inputFileName
    global inputfile_extension

    inputFile = os.path.basename(filepath)
    inputFileName ,inputfile_extension = os.path.splitext(filepath)

    print(inputfile_extension)


    inputFileName = inputFile.replace(str(inputfile_extension),"")

    print(inputFileName)

    global currentTime
    sep = '.'

    currentTime = datetime.datetime.now()
    currentTime = str(currentTime).replace(":","_")
    currentTime = str(currentTime).split(sep, 1)[0]
    currentTime = "__" + currentTime

    print(currentTime)
    logging.basicConfig(filename='Log/'+str(inputFileName)+str(currentTime)+'.log',level=logging.DEBUG) 

    #proccessFileName = inputFileName
    worksheet = workbook.active
    #for i in worksheet.iter_rows(values_only=True):
        #print(i)
    totalSearch = 1
    for column_data in worksheet[skuRow]:
        if column_data.value != None:
            totalSearch += 1
            if column_data.value not in searchIds:
                searchIds.append(column_data.value)
        
    #print (searchIds)
    #remove nonetype from list
    #searchIds = [i for i in searchIds if i is not None]
    #totalSearch = len(searchIds)
    
    print(str(totalSearch) +" Ids Loaded")

    
    # change value
    #worksheet['A1'] = "Luxury Market_1"
    # worksheet['A1'].value = 10
    workbook.save('Input/'+inputFile) 

    return searchIds, filepath ,totalSearch,inputFileName,currentTime,inputfile_extension, True

def search(xf,targetSite):   
            xf = str(xf)
            targetSite = str(targetSite)

            headers = {
                "User-Agent":
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/   101.0.4951.54 Safari/537.36"
            }

            params = {
                "q": xf, # search query
                "tbm": "isch",                # image results
                "hl": "en",                   # language
                "ijn": "0"                    # page number
            }
            
            print("SEARCH_______" + targetSite)
            if targetSite != "":
                logging.debug("SEARCHING TARGET SITE: " + targetSite)
                xf = "site:https://www." + targetSite + ' "' + xf + '"'
                print(xf)
                logging.debug(xf)
            print("Target Site Not in use")
            logging.debug("Target Site Not in use")
            r = requests.get("https://www.google.com/search?q=" + xf +"&newwindow=1&tbm=isch&safe=active",  params=params, headers=headers, timeout=30,
                stream=True)

            #originalXF = xf
            print (r.raw._fp.fp.raw._sock.getpeername())
            global Descrip
            Descrip = LR().get(r.text, '"2008":[null,"', '"]}],null,') 
            #while Descrip == "Error: unset variables, Or substrings aren't in the main string":
            #    print("ok")
            #    xf = xf.rstrip(xf[-1])
            #    nxf = xf + '\"'
            #    print(nxf)
            #    if len(nxf) == 45: 
            #        print("Could not find")
            #        break
            #    r = requests.get("https://www.google.com/search?q=" + nxf +"&newwindow=1&tbm=isch",  params=params, headers=headers, timeout=30,
            #    stream=True)
            #    Descrip = LR().get(r.text, '"2008":[null,"', '"]}],null,') 
            descrip_str = str(Descrip[0].lower())

            #found = re.search(key_in_descrip.replace("\n", "") ,descrip_str)
            #print(found)

            #if found:
            #print("KEYWORD FOUND")
            global soup
            soup = BeautifulSoup(r.text, 'lxml')

            return r,soup,descrip_str
   




def get_original_images(soup):
  
    google_images = []

    all_script_tags = soup.select("script")

    # https://regex101.com/r/48UZhY/4
    matched_images_data = "".join(re.findall(r"AF_initDataCallback\(([^<]+)\);", str(all_script_tags)))

    # https://kodlogs.com/34776/json-decoder-jsondecodeerror-expecting-property-name-enclosed-in-double-quotes
    # if you try to json.loads() without json.dumps() it will throw an error:
    # "Expecting property name enclosed in double quotes"
    matched_images_data_fix = json.dumps(matched_images_data)
    matched_images_data_json = json.loads(matched_images_data_fix)

    # https://regex101.com/r/VPz7f2/1
    matched_google_image_data = re.findall(r'\"b-GRID_STATE0\"(.*)sideChannel:\s?{}}', matched_images_data_json)

    # https://regex101.com/r/NnRg27/1
    matched_google_images_thumbnails = ", ".join(
        re.findall(r'\[\"(https\:\/\/encrypted-tbn0\.gstatic\.com\/images\?.*?)\",\d+,\d+\]',
                   str(matched_google_image_data))).split(", ")

    thumbnails = [
        bytes(bytes(thumbnail, "ascii").decode("unicode-escape"), "ascii").decode("unicode-escape") for thumbnail in matched_google_images_thumbnails
    ]

    # removing previously matched thumbnails for easier full resolution image matches.
    removed_matched_google_images_thumbnails = re.sub(
        r'\[\"(https\:\/\/encrypted-tbn0\.gstatic\.com\/images\?.*?)\",\d+,\d+\]', "", str(matched_google_image_data))

    # https://regex101.com/r/fXjfb1/4
    # https://stackoverflow.com/a/19821774/15164646
    #!NOT SURE
    global matched_google_full_resolution_images
    #!NOT SURE
    matched_google_full_resolution_images = re.findall(r"(?:'|,),\[\"(https:|http.*?)\",\d+,\d+\]", removed_matched_google_images_thumbnails)

    full_res_images = [
        bytes(bytes(img, "ascii").decode("unicode-escape"), "ascii").decode("unicode-escape") for img in matched_google_full_resolution_images
    ]
    return matched_google_full_resolution_images



def small_image(xf):
    r = requests.get("https://www.google.com/search?q=" + xf +"&newwindow=1&tbm=isch&safe=active")
    contentBody = r.content.decode('latin-1')
    SmallImage = LR().get(contentBody, '<div class="NZWO1b"><img class="yWs4tf" alt="" src="', '"')
    
    return SmallImage

from selenium import webdriver 
from selenium.webdriver.chrome.service import Service as ChromeService 
from webdriver_manager.chrome import ChromeDriverManager 


import cv2
import numpy as np
import os
from PIL import Image


def clean_image(filePath):
     
        img = cv2.imread(filePath)
        #cv2.imshow("orig", img)
        #cv2.waitKey(5)
        img = cv2.blur(img,(2,2))
        gray_seg = cv2.Canny(img, 0, 100) # 0, 100
        img = cv2.imread(filePath)
        blurred = cv2.blur(img, (3,3))
        canny = cv2.Canny(blurred, 50, 200) # 50, 200
        ## find the non-zero min-max coords of canny
        pts = np.argwhere(canny>0)
        y1,x1 = pts.min(axis=0)
        y2,x2 = pts.max(axis=0)

        ## crop the region
        cropped = img[y1:y2, x1:x2]

        cv2.imwrite(filePath, cropped)


def download_image_headless(url, image_name):
    """
    Download an image from a URL using a headless browser.
    
    Args:
        url (str): The URL of the image to download.
        image_name (str): The name to save the downloaded image as.
    """
    # Set up a headless browse
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36'
    options = webdriver.ChromeOptions()
# specify headless mode
    options.add_argument('headless')
# specify the desired user agent
    options.add_argument(f'user-agent={user_agent}')
    #try:
    driver = webdriver.Chrome(chrome_options=options)

    driver.get(url)
    logging.debug("Page URL: " + str(driver.current_url) )
    print("Page URL:", driver.current_url) 
    logging.debug("Page Title: " + str(driver.title))
    print("Page Title:", driver.title)
    if driver.title == "Access Denied":
        return None
    driver.save_screenshot(image_name)
    return image_name

def imageDownload(url,newpath,imageName,s,cookies):
    timeout = 15 
    headers = { 
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "en-US,en;q=0.9",
        "cookie": "Cookie: Something",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0"
        }     
    if ".webp" in str(url):
        print("NEED TO CONVERT IMAGE")
        logging.debug("NEED TO CONVERT IMAGE")
        with open(newpath+"/"+imageName + '.webp', 'wb') as handle:
                    response = s.get(url, headers=headers, timeout=timeout,stream=True)
                    if not response.ok:
                        print (response)
                    for block in response.iter_content(1024):
                        if not block:
                            break
                        
                        handle.write(block)
        im = Image.open(newpath+"/"+imageName + '.webp').convert("RGB")
        im.save(newpath+"/"+imageName + '.png','png')
    elif ".mpo" in str(url):
        print("broken image")
        logging.debug("broken image")
    else:
        try:
            with open(newpath+"/"+imageName + '.png', 'wb') as handle:
                    response = s.get(url, headers=headers, timeout=timeout,stream=True)
                    if not response.ok:
                        print (response)
                    for block in response.iter_content(1024):
                        if not block:
                            break
                        
                        handle.write(block)

            ######WEBP DETECTION            
            whatType = imghdr.what(newpath+"/"+imageName + '.png')
            print(whatType)
            ######WEBP DETECTION 
            if whatType == 'webp':
                 logging.debug("Webp image detected")
                 with open(newpath+"/"+imageName + '.webp', 'wb') as handle:
                    response = s.get(url, headers={"User-Agent":     "Mozilla/5.0"}, cookies=cookies,stream=True)
                    if not response.ok:
                        print (response)
                    for block in response.iter_content(1024):
                        if not block:
                            break
                        
                        handle.write(block)
                 im = Image.open(newpath+"/"+imageName + '.webp').convert("RGB")
                 im.save(newpath+"/"+imageName + '.png','png')
                 
        except Exception as exc:
            logging.error("IMAGE DOWNLOAD ERROR:")
            print(exc)
            logging.error(exc)
            print(imageName)
            logging.error(imageName)
            print(url)
            logging.error(url)
            with open('BrokenLinks/'+inputFileName+currentTime+'.txt', 'a',encoding="utf-8") as f:
                    f.write(imageName +"\t"+url+"\n")
            
    return True





def verify_png_image_single(imagepath):
    try:
        img = Image.open(imagepath)
        img.getdata()[0]
    except OSError as osexfc:
        logging.error("IMAGE verify ERROR: %s " % osexfc)
        return False
    return True
###NEED TO ADD FORLOOP TO VERIFY ALL DOWNLOADED IMAGES 
