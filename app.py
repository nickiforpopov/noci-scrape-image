import customtkinter
import os
import requests, csv,openpyxl,time,sys,datetime
from PIL import Image
from tkinter import filedialog
from CTkMessagebox import CTkMessagebox
from LR import LR

from os.path import exists

import speedtest
import public_ip 
import time, threading

from functions import getsku
from functions import search
from functions import get_original_images
from functions import small_image
from functions import verify_png_image_single
from functions import imageDownload
from functions import decode_url
from functions import download_image_headless
from functions import clean_image


import functions
from openpyxl import load_workbook

import subprocess
from PIL import Image
from PIL import JpegImagePlugin 
JpegImagePlugin._getmp = lambda: None   


##! SEARCH SPECIFIC SITE MIGHT NEED DIFFRENT IMAGE DOWNLOAD> THREADING MIGHT CAUSE POOL TO OVERFILL

class App(customtkinter.CTk):
       
    def __init__(self):
        super().__init__()
        fonttypeButton = "System",15
        fonttypeLabel = customtkinter.CTkFont(size=15, weight="bold")
        self.iconbitmap("ui_images/favicon.ico")
        self.title("Icon Image Scrape")
        self.geometry("1000X450")
        customtkinter.set_appearance_mode("dark")
        # set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # load images with light and dark mode image
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "ui_images")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "icon_scrape.ico")), size=(26, 26))

        self.upload_icon_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "upload_light.png")), size=(20, 20))
        self.scrape_icon_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "scrape.png")), size=(20, 20))
        self.clear_icon_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "clear_light.png")), size=(20, 20))

        self.home_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "home_dark.png")),
                                                 dark_image=Image.open(os.path.join(image_path, "home_light.png")), size=(20, 20))
        
        #self.release_notes = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "notes_dark.png")),
                                                # dark_image=Image.open(os.path.join(image_path, "notes_light.png")), size=(20, 20))        

        #self.coming_soon = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "coming_soon_dark.png")),
                                                 #dark_image=Image.open(os.path.join(image_path, "coming_soon_light.png")), size=(20, 20))   
        
        self.chat_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "output_dark.png")),
                                                 dark_image=Image.open(os.path.join(image_path, "output_light.png")), size=(20, 20))
        self.add_user_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(image_path, "instructions_dark.png")),
                                                     dark_image=Image.open(os.path.join(image_path, "instructions_light.png")), size=(20, 20))

        # create navigation frame
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(6, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text=" V4", image=self.logo_image,
                                                             compound="left", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = customtkinter.CTkButton(self.navigation_frame, font=(fonttypeButton),corner_radius=0, height=40, border_spacing=10, text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Output",
                                                      fg_color="transparent", font=(fonttypeButton),text_color=("gray10", "gray90"), hover_color=("gray70",    "gray30"),
                                                      image=self.chat_image, anchor="w", command=self.frame_2_button_event)
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = customtkinter.CTkButton(self.navigation_frame, font=(fonttypeButton),corner_radius=0, height=40, border_spacing=10, text="Guide",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.add_user_image, anchor="w", command=self.frame_3_button_event)
        self.frame_3_button.grid(row=3, column=0, sticky="ew")


        #self.frame_4_button = customtkinter.CTkButton(self.navigation_frame, font=(fonttypeButton),corner_radius=0, height=40, border_spacing=10,text="Release     notes",
                                                      #fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      #image=self.release_notes, anchor="w", command=self.frame_4_button_event)
        #self.frame_4_button.grid(row=4, column=0, sticky="ew")

        ##REPLACE IMAGE
        #self.frame_5_button = customtkinter.CTkButton(self.navigation_frame, font=(fonttypeButton),corner_radius=0, height=40, border_spacing=10, text="Coming Soon",
                                                      #fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      #image=self.coming_soon, anchor="w", command=self.frame_5_button_event)
        #self.frame_5_button.grid(row=5, column=0, sticky="ew")

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["Dark", "Light", "System"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=6, column=0, padx=20, pady=20, sticky="s")

        # create home frame
        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(0, weight=1)

        #self.home_frame_large_image_label = customtkinter.CTkLabel(self.home_frame, text="", image=self.large_test_image)
        #self.home_frame_large_image_label.grid(row=0, column=0, padx=20, pady=10)

        self.home_frame_label_upload = customtkinter.CTkLabel(self.home_frame, font=(fonttypeLabel),text="Upload File:")
        self.home_frame_label_upload.grid(row=1, column=1, padx=20, pady=5,sticky = "w")

        self.home_frame_button_upload = customtkinter.CTkButton(self.home_frame, font=(fonttypeButton),text="Browse", image=self.upload_icon_image,       compound="right", command=self.open_file)
        self.home_frame_button_upload.grid(row=1, column=2, padx=20, pady=10)

        self.home_frame_upload_display = customtkinter.CTkLabel(self.home_frame,text="")
        self.home_frame_upload_display.grid(row=3, column=1, padx=20, pady=10,sticky = "w")



        self.home_frame_label_search = customtkinter.CTkLabel(self.home_frame,  font=(fonttypeLabel),text="Search column: ")
        self.home_frame_label_search.grid(row=4, column=1, padx=20, pady=10,sticky = "w")

        self.home_frame_entry_search = customtkinter.CTkEntry(self.home_frame)
        self.home_frame_entry_search.grid(row=4, column=2, padx=20, pady=10)

        self.home_frame_label_targetSite = customtkinter.CTkLabel(self.home_frame,  font=(fonttypeLabel),text="Search site (optional): ")
        self.home_frame_label_targetSite.grid(row=6, column=1, padx=20, pady=10,sticky = "w")

        self.home_frame_entry_targetSite = customtkinter.CTkEntry(self.home_frame,placeholder_text="sitename.com")
        self.home_frame_entry_targetSite.grid(row=6, column=2, padx=20, pady=10)

        #self.home_frame_search_display = customtkinter.CTkLabel(self.home_frame,text="")
        #self.home_frame_search_display.grid(row=3, column=2, padx=5, pady=5)

        #self.home_frame_label_anchor = customtkinter.CTkLabel(self.home_frame, font=(fonttypeLabel), text="Select anchor column: ")
        #self.home_frame_label_anchor.grid(row=1, column=3, padx=20, pady=10)

        #self.home_frame_entry_anchor = customtkinter.CTkEntry(self.home_frame)
        #self.home_frame_entry_anchor.grid(row=2, column=3, padx=20, pady=10)


        #self.home_frame_anchor_display = customtkinter.CTkLabel(self.home_frame,text="")
        #self.home_frame_anchor_display.grid(row=3, column=3, padx=5, pady=5)

        self.home_frame_button_start = customtkinter.CTkButton(self.home_frame, text="Start",font=(fonttypeButton), image=self.scrape_icon_image,  compound="right", command=self.start_code)
        self.home_frame_button_start.grid(row=8, column=2, padx=20, pady=20)

        self.home_frame_button_clear = customtkinter.CTkButton(self.home_frame, text="Clear",font=(fonttypeButton), image=self.clear_icon_image,   compound="right",  command=self.clear_input)
        self.home_frame_button_clear.grid(row=9, column=2, padx=20, pady=20)

        self.home_frame_label_update = customtkinter.CTkLabel(self.home_frame,  font=(fonttypeLabel),text="")
        self.home_frame_label_update.grid(row=10, column=1, padx=20, pady=10,sticky = "w")

        self.progressbar = customtkinter.CTkProgressBar(self.home_frame)
        self.progressbar.grid(row=11, column=1, padx=20, pady=10,sticky = "w")
        self.progressbar.configure(mode="indeterminate_speed ")
        self.progressbar.set(0)

        self.home_frame_label_time = customtkinter.CTkLabel(self.home_frame,  font=(fonttypeLabel),text="")
        self.home_frame_label_time.grid(row=12, column=1, padx=20, pady=10,sticky = "w")


        # create second frame
        self.second_frame = customtkinter.CTkScrollableFrame(self, corner_radius=0, width=600 ,height=450,fg_color="transparent")
        self.second_frame.grid_columnconfigure(0, weight=1)

        
        # create third frame
        self.third_frame = customtkinter.CTkFrame(self, corner_radius=0,fg_color="transparent")
        self.third_frame.grid_columnconfigure(0, weight=1)
        
        self.third_frame_header = customtkinter.CTkLabel(self.third_frame, font=customtkinter.CTkFont(size=25, weight="bold"),text="Instructions")
        self.third_frame_header.grid(row=1, column=1, padx=20, pady=15,sticky = "w")    


        self.third_frame_content = customtkinter.CTkLabel(self.third_frame, font=(fonttypeButton),text="Icon Image Scrape")
        self.third_frame_content.grid(row=2, column=1, padx=20, pady=5,sticky = "w")

        self.third_frame_content_1 = customtkinter.CTkLabel(self.third_frame, font=(fonttypeButton),text="1. Image column height must be at least 100")
        self.third_frame_content_1.grid(row=3, column=1, padx=20, pady=5,sticky = "w")        

        self.third_frame_content_2 = customtkinter.CTkLabel(self.third_frame, font=(fonttypeButton),text="2. Remove all slashes from search column")
        self.third_frame_content_2.grid(row=4, column=1, padx=20, pady=5,sticky = "w")   

        #self.third_frame_content_3 = customtkinter.CTkLabel(self.third_frame, font=(fonttypeButton),text="3. Rows should not exceed 700")
        #self.third_frame_content_3.grid(row=5, column=1, padx=20, pady=5,sticky = "w")   

        self.third_frame_content_4 = customtkinter.CTkLabel(self.third_frame, font=(fonttypeButton),text="4. Only supports one sheet")
        self.third_frame_content_4.grid(row=6, column=1, padx=20, pady=5,sticky = "w")  

        self.third_frame_content_5 = customtkinter.CTkLabel(self.third_frame, font=(fonttypeButton),text="5. Image will be added to column A")
        self.third_frame_content_5.grid(row=7, column=1, padx=20, pady=5,sticky = "w")  




        # create fourth frame
        #self.fourth_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        #self.fourth_frame.grid_columnconfigure(0, weight=1)
        #
        #self.fourth_frame_header = customtkinter.CTkLabel(self.fourth_frame, font=customtkinter.CTkFont(size=25, weight="bold"),text="Release Notes")
        #self.fourth_frame_header.grid(row=1, column=1, padx=20, pady=15)    
        #self.fourth_frame_content = customtkinter.CTkLabel(self.fourth_frame, font=(fonttypeButton),text="Icon Image Scrape V3-BETA:\n\n3-30-2023\n\n1.Removed     #Anchor  Input, for simplicity, will add back in in the future on advanced search tab. Image will always be added to column A\n\n 2. DownloadSpeed  #Improvments\n\n 3.    Broken Images will be skipped\n\n 4. Release Notes Tab added\n\n 5. Coming Soon Tab added.")
        #self.fourth_frame_content.grid(row=2, column=1, padx=20, pady=5)  
#
        ## create fifth frame
        #self.fifth_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        #self.fifth_frame.grid_columnconfigure(0, weight=1)
        #
        #self.fifth_frame_header = customtkinter.CTkLabel(self.fifth_frame, font=customtkinter.CTkFont(size=25, weight="bold"),text="Coming Soon")
        #self.fifth_frame_header.grid(row=1, column=1, padx=20, pady=15)    
        #self.fifth_frame_content = customtkinter.CTkLabel(self.fifth_frame, font=(fonttypeButton),text="1. Advanced search tab(Ability to choose where imagewill   #sit,  select specific site)\n\n 2. Progress Bar \n\n 3. Whitelist sites to pull from, all other images will need users review\n\n 4. Insert ImageSpeed   #Updates\n\n    5. Error Reporting and Logging for users\n\n6. Auto IP Switch")
        #self.fifth_frame_content.grid(row=2, column=1, padx=20, pady=5)


        # select default frame
        self.select_frame_by_name("home")

    def open_finished_file(self, filename):
        # Check if file exists
        functions.logging.debug("Filename:  %s " % filename)
        print(filename)
        filepath_open = os.path.realpath(filename)
        filepath_open= str(filepath_open).replace(str(filename),"")
        filepath_open = filepath_open + "Output/" + str(filename)
        functions.logging.debug("filepath_open:  %s " % filepath_open)
        print(filepath_open)
        if not os.path.exists(filepath_open):
            print("File does not exist!")
            functions.logging.debug("File does not exist!")
            return
    
        # Open file with the default program for its file type
        try:
            subprocess.Popen([filepath_open], shell=True)
        except Exception as e:
            print(f"Error opening file: {e}")
            functions.logging.debug(f"Error opening file: {e}")
    def open_finished_file_box(self, filename):
        # Check if file exists
        functions.logging.debug("Filename:  %s " % filename)
        print(filename)
        filepath_open = os.path.realpath(filename)
        filepath_open= str(filepath_open).replace(str(filename),"")
        functions.logging.debug("filepath_open:  %s " % filepath_open)
        print(filepath_open)
        if not os.path.exists(filepath_open):
            print("File does not exist!")
            functions.logging.debug("File does not exist!")
            return
    
        # Open file with the default program for its file type
        try:
            subprocess.Popen([filepath_open], shell=True)
        except Exception as e:
            print(f"Error opening file: {e}")
            functions.logging.debug(f"Error opening file: {e}")

    def get_speed(self):
  
        try:
            stest = speedtest.Speedtest()

        except:
            msg = CTkMessagebox(title="Internet is too slow!", message="Cannot Calculate Download Speed \nReset VPN and Select Retry" ,
            icon="cancel", option_1="Exit", option_2="Retry")
            response = msg.get()
            if response=="Exit":
                return False
            if response == "Retry":
                stest = speedtest.Speedtest()

        internetSpeed = stest.download()  

        internetSpeed = internetSpeed /1000000
        internetSpeed =float(internetSpeed)
        internetSpeed = round(internetSpeed,2)  
        return internetSpeed 
    def slow_connection_caught(self):  
          
        self.home_frame_label_update.configure(text="CONNECTION SEEMS TO BE SLOW")
        app.update()

        msg2 = CTkMessagebox(title="CONNECTION SEEMS TO BE SLOW", message="Please Select Option Below",
           icon="warning", option_1="Exit",  option_2="Continue", option_3="Recover")
        response = msg2.get()
        if response=="Exit":
            return False

        if response=="Recover":
            
            self.home_frame_label_update.configure(text="Testing Internet Speed...Please Wait...")
            app.update()
            internetSpeed = self.get_speed()
            if internetSpeed == False:                                        
               return False
            while internetSpeed < 23:
                self.home_frame_label_update.configure(text="Testing Internet Speed...Please Wait...")
                app.update()
                internetSpeed = self.get_speed()
                if internetSpeed == False:                                        
                    return False
                if internetSpeed > 23:
                  break
                # Show some retry/cancel warnings
                msg4 = CTkMessagebox(title="Internet is too slow!", message="Download Speed is " + str(internetSpeed) + " Mbps  \nReset VPN and Select Retry" ,
                icon="info", option_1="Cancel", option_2="Retry")
                response = msg4.get()
                if response=="Cancel":
    ############! NEED TO FIX CANCEL
                    return False
            
            msg4 = CTkMessagebox(title="Speed test passed!", message="Download Speed is fine\nImage servers are slow\nPlease be patient..." ,
            icon="check", option_1="Ok")


    def speed_on_start(self):  
        self.home_frame_label_update.configure(text="Verifying Connection...")
        app.update()
        internetSpeed = self.get_speed()
        if internetSpeed == False:                                        
            return False  
        
        while internetSpeed < 23:
            self.home_frame_label_update.configure(text="Testing Internet Speed...Please Wait...")
            app.update()
            internetSpeed = self.get_speed()
            if internetSpeed == False:                                        
                return False
            if internetSpeed > 23:
                break
                # Show some retry/cancel warnings   
            msg3 = CTkMessagebox(title="Internet is too slow!", message="Download Speed is " + str(internetSpeed) + " Mbps  \nReset VPN and Select Retry" ,
            icon="info", option_1="Cancel", option_2="Retry")
            response = msg3.get()
            if response=="Cancel":
    ############!CANCEL??
                return False
            
        self.home_frame_label_update.configure(text="Speed Test Passed")
        app.update()

    def change_ip(self):
        self.home_frame_label_update.configure(text="IP NEEDS TO BE CHANGED!")
        app.update()
        old_ip= public_ip.get()
        verify = False
        while verify == False:
            msg2 = CTkMessagebox(title="IP NEEDS TO BE CHANGED!", message="IP: " + str(old_ip) + " \nis about to be blocked\nPlease Reset VPN...",
               icon="warning", option_1="Verify New IP")
            response = msg2.get()
            if response=="Verify New IP":
                new_ip = public_ip.get()
                print(str(new_ip) + "---------------------------------------------------------")

            if new_ip != old_ip and new_ip != "216.158.128.106":
                self.home_frame_label_update.configure(text="IP SUCCESSFULLY CHANGED")
                app.update()
                verify = True
        
            

    def display_files(self):
    
        # Clear any existing widgets in self.second_frame
        for widget in self.second_frame.winfo_children():
            widget.destroy()
        self.second_frame_header = customtkinter.CTkLabel(self.second_frame, font=customtkinter.CTkFont(size=25, weight="bold"),text="Complete")
        self.second_frame_header.grid(row=1, column=1, padx=20, pady=15,sticky = "w")  

        # Get a list of all the files in the Output/ directory
        file_list = os.listdir("Output/")

        # Add a CTkLabel and CTkButton for each file in the grid
        for ind,filename in enumerate(file_list):
            label = customtkinter.CTkLabel(self.second_frame, text=filename,font = customtkinter.CTkFont(size=12))
            label.grid(row=ind+2,column=2, padx=5,pady=10,sticky = "w")
            button = customtkinter.CTkButton(self.second_frame,font=("System",15), text="Open", command=lambda f=filename: self.open_finished_file(f))
            button.grid(row=ind+2,column=1, padx=5,pady=10,sticky = "w")


    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")
        self.frame_3_button.configure(fg_color=("gray75", "gray25") if name == "frame_3" else "transparent")
        #self.frame_4_button.configure(fg_color=("gray75", "gray25") if name == "frame_4" else "transparent")
        #self.frame_5_button.configure(fg_color=("gray75", "gray25") if name == "frame_5" else "transparent")

        # show selected frame
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
            self.display_files()
        else:
            self.second_frame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()
        #if name == "frame_4":
        #    self.fourth_frame.grid(row=0, column=1, sticky="nsew")
        #else:
        #    self.fourth_frame.grid_forget()
        #if name == "frame_5":
        #    self.fifth_frame.grid(row=0, column=1, sticky="nsew")
        #else:
            #self.fifth_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def frame_3_button_event(self):
        self.select_frame_by_name("frame_3")

    #def frame_4_button_event(self):
       # self.select_frame_by_name("frame_4")
    #def frame_5_button_event(self):
        #self.select_frame_by_name("frame_5")    

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)
    def open_file(self):
        # open file explorer and save file path
        self.file_path = filedialog.askopenfilename()
        # display selected file path
        if self.file_path:
            self.home_frame_upload_display.configure(text=self.file_path)


    def check_inputs(self):
        # check if a file has been selected
        if not self.file_path:
            CTkMessagebox(title="Error", message="File Not Selected", icon="cancel")
            return False
        
        if not self.home_frame_entry_search.get().isalpha():
            CTkMessagebox(title="Error", message="Enter Column Letter", icon="cancel")
            return False
        
            #app.destroy() 
        #if not self.home_frame_entry_anchor.get().isalpha():
        #    CTkMessagebox(title="Error", message="Please enter Column Letter", icon="cancel")
        #    return False

        else:
            self.ask_question()
            return True  
    def clear_input(self):
        self.home_frame_upload_display.configure(text="")
        self.home_frame_label_update.configure(text="")
        self.home_frame_label_time.configure(text="")
        self.home_frame_label_update.configure(text="")
        #self.home_frame_entry_anchor.delete(0, 'end')
        self.home_frame_entry_search.delete(0,'end' )
        self.home_frame_entry_targetSite.delete(0,'end' )

    def ask_question(self):
        try:  
        # get yes/no answers
            msg = CTkMessagebox(title="Continue?", message="Do you want to start the program?\nFile: "+ str(self.file_path)+"\nSearch Column: "  +str(self.home_frame_entry_search.get()),
                               icon="question", option_1="No",  option_2="Yes")
            response = msg.get()
 
            if response=="Yes":
                destroy = False
                start_time = time.time()
                log_time = datetime.datetime.now()
                self.search_col = self.home_frame_entry_search.get()
                self.targetSite = self.home_frame_entry_targetSite.get()
                self.progressbar.start()
                #self.anchor_col = self.home_frame_entry_anchor.get()


                global filepath
                global skuRow
                global targetSite
                filepath = self.file_path
                skuRow = self.search_col
                targetSite = self.targetSite
                print(targetSite)
                #search.anchor = self.anchor_col


                print("Input Variables:")
                print(f"  filepath: {self.file_path}")
                print(f"  skuRow: {self.search_col}")
                print(f"  targetSite: {self.targetSite}")
                #print(f"  search_col: {self.anchor_col}")

                ### START OF LOGIC 
                getskuResponse = getsku(filepath,skuRow)
                totalSkus = getskuResponse[2]
                currentTime = getskuResponse[4]  
                inputFileName = getskuResponse[3]
                inputfile_extension = getskuResponse[5]

                functions.logging.debug("Proccess Started: %s " % log_time)
                functions.logging.debug("Input Variables:")
                functions.logging.debug(f"  file_path: {self.file_path}")
                functions.logging.debug(f"  search_col: {self.search_col}")


                self.home_frame_label_update.configure(text="Checking IP...Please Wait....")
                app.update()  


                ip_add= public_ip.get()
  
                print(str(ip_add) + "---------------------------------------------------------")
                if str(ip_add) == "216.158.128.106":
                    CTkMessagebox(title="Error", message="VPN IS NOT CONNECTED", icon="cancel")
                    self.home_frame_label_update.configure(text="Please, connect vpn....") 
                    functions.logging.debug("PROGRAM KILLED, VPN NOT CONNECTED")
                    app.update() 

                else:
                    check_speed_first = self.speed_on_start()

                    if check_speed_first == False:
                        destroy = True
                        app.destroy()

                    if getskuResponse[6] == True:
                        self.home_frame_label_update.configure(text=str(totalSkus) + " Search ID's Loaded....")
                        app.update()


                    global xf
                    global index






                    searchstart_time = datetime.datetime.now()
                    functions.logging.debug("Search Started: %s " % searchstart_time)
                    firstTenElapsed = 0
                    counter_index = 0
                    for index,xf in enumerate(getskuResponse[0]):
                        xf = str(xf)
                        targetSite = str(targetSite)
                        global status    
                        status = "Searching... " + str(index+1) + " out of " + str(totalSkus)

                        searchStart_elap = time.time()

                        searchResponse = search(xf,targetSite)

                        if "Looks like there aren’t any matches for your search" in searchResponse[0]:
                          with open('DontExist.txt', 'a') as f:
                            f.write(xf+"\n")

                        elif "Looks like there aren’t any matches for your search" not in searchResponse[0]:
                            hiQResponse = get_original_images(searchResponse[1])

                        print(hiQResponse)
                        functions.logging.debug("High Quality Response: %s " % hiQResponse)

                        #if len(hiQResponse) == 0:  
                            #smQResponse = small_image(xf)
                            #print(smQResponse) #NEED TO WRITE TO FILE
                            #with open('Data/'+inputFileName+currentTime+'.txt', 'a',encoding="utf-8") as f:
                                #f.write(xf.replace("\n", "")+"\t"+smQResponse+"\t"+searchResponse[2]+"\n")
                        #else:
                        if len(hiQResponse) == 0:
                            continue
                        if '(' in hiQResponse[0]:
                                        functions.logging.debug("Started Url Cleaning: %s " % hiQResponse[0])
                                        MessyPart = LR().get(hiQResponse[0], '(', ')') 
                                        #print(MessyPart) 
                                        cleanUrl = hiQResponse[0].replace('(' + str(MessyPart[0] + ')'), "")
                                        #print(cleanUrl)

                                        with open('Data/'+inputFileName+currentTime+'.txt', 'a',encoding="utf-8") as f:
                                            functions.logging.debug("Encoded: %s " % cleanUrl)
                                            cleanUrl = decode_url(cleanUrl)
                                            functions.logging.debug("Url Cleaning / Decoding Successful: %s " % cleanUrl)
                                            f.write(xf.replace("\n", "")+"\t"+cleanUrl+"\t"+searchResponse[2]+"\n")
                        else:
                                #!NOT SURE


                                with open('Data/'+inputFileName+currentTime+'.txt', 'a',encoding="utf-8") as f:
                                    functions.logging.debug("Encoded: %s " % hiQResponse[0])
                                    decodedUrl = decode_url(str(hiQResponse[0]))
                                    functions.logging.debug("Url Cleaning / Decoding Successful: %s " % decodedUrl)
                                    f.write(xf.replace("\n", "")+"\t"+decodedUrl+"\t"+searchResponse[2]+"\n")

                        self.home_frame_label_update.configure(text=status)
                        app.update()       
                        searchElapsed = time.time() - searchStart_elap
                        fsearchElapsed = float(searchElapsed)
                        print(fsearchElapsed)
                        formatsearchElapsed = format(fsearchElapsed,".2f")

                        firstTenElapsed = float(firstTenElapsed) + float(fsearchElapsed)
                        print("ELAPSED ____________________________" + str(firstTenElapsed))
                        
                        if counter_index == 10 and firstTenElapsed > 19:
                            responseConnection = self.slow_connection_caught()
                            if responseConnection == False:
                                destroy = True
                                app.destroy()

                        if counter_index == 500:
                            firstTenElapsed = 0
                            counter_index = 0
                            self.change_ip()

                        print("Index" + str(index))
                        functions.logging.debug("Search Request Elapsed Time:  %s " % formatsearchElapsed)
                        counter_index = counter_index + 1
                        print("Index Counter" + str(counter_index))


                    #!getskuResponse[4] = current time AND inputfilename getskuResponse[3]
                    #########DOWNLOAD IMAGES
                    downstart_time = datetime.datetime.now()
                    functions.logging.debug("Download Started: %s " % downstart_time)
                    s = requests.Session()
                    cookies = dict(BCPermissionLevel='PERSONAL')
                    newpath = 'Images/'+inputFileName+currentTime
                    if not os.path.exists(newpath):
                        os.makedirs(newpath)


                    with open('Data/'+inputFileName+currentTime+'.txt',encoding="utf8") as f:
                        txt_reader = csv.reader(f, delimiter="\t")
                        # Skip the first row, which is the header
                        next(txt_reader)


                        t1 = time.time()

                        threads = []

                        for iter,rowD in enumerate(txt_reader):

                            downStart_elap = time.time()

                            self.home_frame_label_update.configure(text="Downloading image " +str(iter+1) +" out of " + str(totalSkus))
                            app.update()  
                            (imageName,url, decs) = rowD

                            temp = threading.Thread(target=imageDownload, args=(url,newpath,imageName,s,cookies))
                            temp.start()
                            threads.append(temp)
                            
                            downElapsed = time.time() - downStart_elap
                            fdownElapsed = float(downElapsed)
                            formatdownElapsed = format(fdownElapsed,".2f")
                            functions.logging.debug("Download Request Elapsed Time:  %s " % formatdownElapsed)

                        for thread in threads:
                            thread.join()

                        t2 = time.time()
                        timeDif = t2-t1
                        print("Time takes Download with Threading : " + str(timeDif))
                        functions.logging.debug("Time takes Download with Threading :   %s " % timeDif)
                    #########DOWNLOAD IMAGES END                

                    ###NEED TO ADD FORLOOP TO VERIFY ALL DOWNLOADED IMAGES 
                    
                    ###INSERT IMAGE START
                    
                    broken_file_exists = exists('BrokenLinks/'+inputFileName+currentTime+'.txt')
                    print(broken_file_exists)
                    recovAfter = False
                    if broken_file_exists == True:
                        recovAfter = True
                    
                    
                        
                    ###INSERT IMAGE START
                    insertstart_time = datetime.datetime.now()
                    functions.logging.debug("Attaching Started: %s " % insertstart_time)
                    self.home_frame_label_update.configure(text="Attaching Images...")
                    app.update() 



                    print(newpath)
                    functions.logging.debug("newpath: %s " % newpath)
                    workbook = load_workbook(filepath)
                    worksheet = workbook.active
                    interation = 0
                    for interDex, column_data in enumerate(worksheet[skuRow]):

                        print("start image proccessing =====================================")
                        functions.logging.debug("start image proccessing =====================================")

                        print(interDex)
                        functions.logging.debug("interDex: %s " % interDex)
                        print(column_data.value)
                        functions.logging.debug("column_data.value: %s " % column_data.value)

                        imagepath = newpath+'/'+str(column_data.value)+".png"
                        check_file = os.path.isfile(imagepath)
                        print(os.path.isfile(imagepath))
                        functions.logging.debug("os.path.isfile(imagepath): %s " % os.path.isfile(imagepath))
                        if check_file == True:
                            verifyImg = verify_png_image_single(imagepath)
                            print("IMAGE VERIFY: "+ str(verifyImg))
                            functions.logging.debug("IMAGE VERIFY: "+ str(verifyImg))
                            if verifyImg == False:
                                print("IMAGE ERROR ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                functions.logging.debug("IMAGE ERROR ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                            else:
                                imageSize = os.path.getsize(imagepath)
                                print("IMAGE SIZE "+str(imageSize))
                                functions.logging.debug("IMAGE SIZE "+str(imageSize))
                                if imageSize < 3000:
                                    print("IMAGE FILE IS BROKEN")
                                    functions.logging.debug("IMAGE FILE IS BROKEN")
                                else:
                                    interation = interation + 1
                                    self.home_frame_label_update.configure(text="Image " +str(interation) + " of " + str(totalSkus)  + " Attached")
                                    app.update() 
                                    imgOpen = openpyxl.drawing.image.Image(imagepath)
                                    #size down
                                    imgPil = Image.open(imagepath)
                                    w, h = imgPil.size
                                    nw = 100
                                    nh = (h/w)*nw
                                    imgOpen.height = nh
                                    imgOpen.width = nw
                                    #insert index is i from iteration of ARRAY
                                    imgOpen.anchor = "A"+str(interDex+1)
                                    worksheet.add_image(imgOpen)


                        print("end image proccessing")
                        functions.logging.debug("end image proccessing")
                    if recovAfter != True:    
                        workbook.save('Output/'+inputFileName+ currentTime+"_DONE_"+inputfile_extension)
                        openRecentFile = r'Output/'+inputFileName+ currentTime+"_DONE_"+inputfile_extension
                    else: 
                        workbook.save('Output/'+inputFileName+ currentTime+"_TEMP_"+inputfile_extension)
                        openRecentFile = r'Output/'+inputFileName+ currentTime+"_TEMP_"+inputfile_extension

                    if recovAfter == True:        

                            iElapedTime = time.time() - start_time
                            fElapedTime = float(iElapedTime)
                            fElapedTime = round(fElapedTime,0)
                            ElapedTimeMin = fElapedTime / 60
                            ElapedTimeMin = round(ElapedTimeMin,2)
                            ElapedTimeMin = str(ElapedTimeMin).replace(".",":" )
                            #formatElapsedTime = format(fElapedTime,".2f")
                            self.home_frame_label_time.configure(text="Quick Search Elapsed Time --- %s minutes" % ElapedTimeMin)
                            app.update()         
                            self.home_frame_label_update.configure(text="Quick Scan Complete, View Output")
                            app.update() 

                            #!MSG BOX QUIUCK COMPLETE
                            #msg5 = CTkMessagebox(title="Complete", message="Quick Scan Complete\nPlease wait for remaining images to download...",
                            #icon="check", option_1="Ok", option_2="Open File")
                            #response = msg5.get()
                            #if response=="Open File":
                            #self.open_finished_file_box(openRecentFile)
                            #!MSG BOX QUIUCK COMPLETE
                            self.open_finished_file_box(openRecentFile)


                            with open('BrokenLinks/'+inputFileName+currentTime+'.txt',encoding="utf8") as fF:
                                row_count = len(fF.readlines())
                            with open('BrokenLinks/'+inputFileName+currentTime+'.txt',encoding="utf8") as f:
                                txt_reader = csv.reader(f, delimiter="\t")
                                #row_count = sum(1 for row in txt_reader)
                                print(row_count)
                                # Skip the first row, which is the header
                                next(txt_reader)

                                for iters,rowA in enumerate(txt_reader):
                                    self.home_frame_label_update.configure(text="Downloading " +str(iters+1) +" out of " + str(row_count) + " remaining images")
                                    app.update()  
                                    (imageName,url) = rowA
                                    brokenPath = newpath+"/"+imageName + '.png'
                                    try:
                                        downResponse = download_image_headless(url,brokenPath)
                                    except:
                                        functions.logging.debug("Error on function call")
                                        print("Error on function call")
                                    functions.logging.debug("Attempting to Save Images Again")
                                    print(downResponse)
                                    if downResponse == None:
                                        continue 
                                    self.home_frame_label_update.configure(text="Recovered!")
                                    app.update() 
                                    try:
                                        clean_image(downResponse)
                                    except:
                                        functions.logging.debug("Error on image crop")
                                        print("Error on image crop")
                            insertstart_time = datetime.datetime.now()
                            functions.logging.debug("Recovery Attaching Started: %s " % insertstart_time)
                            self.home_frame_label_update.configure(text="Attaching Images...")
                            app.update() 
        
        
        
                            print(newpath)
                            functions.logging.debug("newpath: %s " % newpath)
                            workbook = load_workbook(filepath)
                            worksheet = workbook.active
                            interation = 0
                            for interDex, column_data in enumerate(worksheet[skuRow]):
                            
                                print("start image proccessing =====================================")
                                functions.logging.debug("start image proccessing =====================================")
        
                                print(interDex)
                                functions.logging.debug("interDex: %s " % interDex)
                                print(column_data.value)
                                functions.logging.debug("column_data.value: %s " % column_data.value)
        
                                imagepath = newpath+'/'+str(column_data.value)+".png"
                                check_file = os.path.isfile(imagepath)
                                print(os.path.isfile(imagepath))
                                functions.logging.debug("os.path.isfile(imagepath): %s " % os.path.isfile(imagepath))
                                if check_file == True:
                                    verifyImg = verify_png_image_single(imagepath)
                                    print("IMAGE VERIFY: "+ str(verifyImg))
                                    functions.logging.debug("IMAGE VERIFY: "+ str(verifyImg))
                                    if verifyImg == False:
                                        print("IMAGE ERROR ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                        functions.logging.debug("IMAGE ERROR        ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                    else:
                                        imageSize = os.path.getsize(imagepath)
                                        print("IMAGE SIZE "+str(imageSize))
                                        functions.logging.debug("IMAGE SIZE "+str(imageSize))
                                        if imageSize < 3000:
                                            print("IMAGE FILE IS BROKEN")
                                            functions.logging.debug("IMAGE FILE IS BROKEN")
                                        else:
                                            interation = interation + 1
                                            self.home_frame_label_update.configure(text="Image " +str(interation) + " of " + str(totalSkus)  + " Attached")
                                            app.update() 
                                            imgOpen = openpyxl.drawing.image.Image(imagepath)
                                            #size down
                                            imgPil = Image.open(imagepath)
                                            w, h = imgPil.size
                                            nw = 100
                                            nh = (h/w)*nw
                                            imgOpen.height = nh
                                            imgOpen.width = nw
                                            #insert index is i from iteration of ARRAY
                                            imgOpen.anchor = "A"+str(interDex+1)
                                            worksheet.add_image(imgOpen)
        
        
                                print("end image proccessing")
                                functions.logging.debug("end image proccessing")
                            workbook.save('Output/'+inputFileName+ currentTime+"_DONE_"+inputfile_extension)
                            openRecentFile = r'Output/'+inputFileName+ currentTime+"_DONE_"+inputfile_extension


                    iElapedTime = time.time() - start_time
                    fElapedTime = float(iElapedTime)
                    fElapedTime = round(fElapedTime,0)
                    ElapedTimeMin = fElapedTime / 60
                    ElapedTimeMin = round(ElapedTimeMin,2)
                    ElapedTimeMin = str(ElapedTimeMin).replace(".",":" )
                    #formatElapsedTime = format(fElapedTime,".2f")
                    self.home_frame_label_time.configure(text="Total Elapsed Time --- %s minutes" % ElapedTimeMin)
                    app.update()         
                    self.home_frame_label_update.configure(text="Complete, View Output")
                    app.update() 

                    self.progressbar.stop()
                    msg7 = CTkMessagebox(title="Success",message="File Complete, Please View Output Folder",
                    icon="check", option_1="Ok", option_2="Open File")
                    response = msg7.get()
                    if response=="Open File":
                        self.open_finished_file_box(openRecentFile)





        except Exception as exc:
                 if destroy != True:
                    CTkMessagebox(title="Report", message="AN ERROR OCCURED. PLEASE REPORT THIS TO NIK AND TRY AGAIN", icon="cancel")
                    functions.logging.error("APPLICATION ERROR ----- %s" % exc)
                    print(exc)

    def show_checkmark(self):
      # Show some positive message with the checkmark icon
        msg = CTkMessagebox(title="Success",message="File Complete, Please View Output Folder",
                  icon="check", option_1="OK")

        

    def start_code(self):
        if not self.check_inputs():
            return     
    
    
if __name__ == "__main__":
    app = App()
    app.mainloop()

